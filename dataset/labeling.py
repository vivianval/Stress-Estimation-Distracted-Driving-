#!/usr/bin/env python3
import argparse, json, shlex, subprocess, re
from pathlib import Path
from typing import Optional, List, Tuple

import numpy as np
import pandas as pd
from io import BytesIO
from openpyxl import load_workbook

# ---------------- Config ----------------
ROOT = Path("/media/storage/vivib/Structured Study Data")
OUT_DIR = Path("/home/vivib/emoca/emoca/dataset/LABELS_PERFRAME")
SUBJ_RE = re.compile(r"^T\d{3,}$", re.IGNORECASE)
MD_RE   = re.compile(r"(\d+)\s*MD$", re.IGNORECASE)  # "7MD", "7 MD"
# ----------------------------------------

def find_subject_dirs(root: Path) -> List[Path]:
    return sorted([d for d in root.iterdir() if d.is_dir() and SUBJ_RE.match(d.name)],
                  key=lambda p: p.name)

def detect_md_folder(subject_dir: Path) -> Optional[Path]:
    for d in subject_dir.iterdir():
        if d.is_dir() and MD_RE.search(d.name.replace("_","").replace("-","")):
            return d
    return None

def detect_md_drive_num(folder: Path) -> Optional[int]:
    m = MD_RE.search(folder.name.replace("_","").replace("-",""))
    return int(m.group(1)) if m else None

def pick_md_video(md_folder: Path, subject_id: str, md_drive: int) -> Optional[Path]:
    """Require a video whose name contains <subject>-<drive:03d> and ends with avi1.avi."""
    tag = f"{subject_id}-{'007'}".lower()
    cand = [p for p in md_folder.glob("*avi1.avi") if tag in p.name.lower()]
    if not cand:
        return None
    cand.sort(key=lambda p: p.name.lower())
    return cand[0]

def ffprobe_frame_times(video_path: Path) -> Optional[np.ndarray]:
    try:
        cmd = ("ffprobe -v error -select_streams v:0 "
               "-show_entries frame=best_effort_timestamp_time "
               "-of json " + shlex.quote(str(video_path)))
        out = subprocess.check_output(cmd, shell=True, text=True)
        data = json.loads(out)
        frames = data.get("frames", [])
        ts = []
        for fr in frames:
            t = fr.get("best_effort_timestamp_time")
            if t is not None:
                try: ts.append(float(t))
                except Exception: pass
        return np.array(ts, dtype=float) if ts else None
    except Exception:
        return None

def find_stm_file(md_folder: Path, subject_id: str, md_drive: int) -> Optional[Path]:
    """Return an STM path (.stm.xlsx or .stm). No writes (handles read-only FS)."""
    pref_xlsx = md_folder / f"{subject_id}-{md_drive:03d}.stm.xlsx"
    if pref_xlsx.exists():
        return pref_xlsx
    pref_stm = md_folder / f"{subject_id}-{md_drive:03d}.stm"
    if pref_stm.exists():
        return pref_stm
    cand = (list(md_folder.glob("*.stm.xlsx")) + list(md_folder.glob("*.STM.xlsx")) +
            list(md_folder.glob("*.stm")) + list(md_folder.glob("*.STM")))
    if not cand:
        return None
    key = f"{subject_id}-{md_drive:03d}".lower()
    cand.sort(key=lambda p: (key not in p.name.lower(), p.name.lower()))
    return cand[0]

def load_workbook_any_suffix(path: Path):
    """Open an Excel workbook even if extension is .stm by reading bytes."""
    with path.open("rb") as f:
        data = f.read()
    return load_workbook(filename=BytesIO(data), data_only=True)

def read_intervals_seconds(stm_path: Path) -> List[Tuple[float, float]]:
    """
    Robust reader:
    - scans sheet for 'StartTime' and 'EndTime' headers (anywhere)
    - reads numeric values below each header until blank
    """
    wb = load_workbook_any_suffix(stm_path)
    ws = wb.active

    def find_col(header_names: List[str]) -> Optional[int]:
        keys = {n.lower() for n in header_names}
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                v = cell.value
                if isinstance(v, str) and v.strip().lower() in keys:
                    return cell.column
        return None

    col_start = find_col(["starttime", "start time", "start"])
    col_end   = find_col(["endtime", "end time", "end"])
    if col_start is None or col_end is None:
        raise ValueError(f"{stm_path.name}: couldn't locate StartTime/EndTime headers")

    def header_row(col_idx: int) -> Optional[int]:
        keys = {"starttime","start time","start","endtime","end time","end"}
        for r in range(1, ws.max_row+1):
            v = ws.cell(row=r, column=col_idx).value
            if isinstance(v, str) and v.strip().lower() in keys:
                return r
        return None

    r0 = header_row(col_start); r1 = header_row(col_end)
    if r0 is None or r1 is None:
        raise ValueError(f"{stm_path.name}: header rows not found")

    starts, ends = [], []
    r = r0 + 1
    while r <= ws.max_row:
        v = ws.cell(row=r, column=col_start).value
        if v in (None, ""): break
        try: starts.append(float(v))
        except Exception: break
        r += 1

    r = r1 + 1
    while r <= ws.max_row:
        v = ws.cell(row=r, column=col_end).value
        if v in (None, ""): break
        try: ends.append(float(v))
        except Exception: break
        r += 1

    n = min(len(starts), len(ends))
    return [(s, e) for s, e in zip(starts[:n], ends[:n]) if e > s]

def label_frames_from_intervals(frame_times: np.ndarray, intervals: List[Tuple[float,float]]) -> np.ndarray:
    lab = np.zeros((len(frame_times),), dtype=np.int8)
    if not intervals:
        return lab
    ft = frame_times
    for a, b in intervals:
        if b <= a: continue
        m = (ft >= a) & (ft <= b)
        lab[m] = 1
    return lab

def process_subject(subject_dir: Path, out_dir: Path):
    sid = subject_dir.name
    md_folder = detect_md_folder(subject_dir)
    if not md_folder:
        print(f"[WARN] {sid}: no MD folder -> skip"); return
    md_drive = detect_md_drive_num(md_folder)
    if md_drive is None:
        print(f"[WARN] {sid}: cannot parse MD drive -> skip"); return

    # outputs (resume-safe)
    out_csv  = out_dir / f"{sid}_MD{md_drive}_labels_PERFRAME.csv"
    meta_json= out_dir / f"{sid}_MD{md_drive}_labels_meta.json"
    if out_csv.exists() and meta_json.exists():
        print(f"[SKIP] {sid}: outputs exist -> {out_csv.name}"); return

    video = pick_md_video(md_folder, sid, md_drive)
    if not video:
        print(f"[WARN] {sid}: no video matching {sid}-{md_drive:03d}*.avi1.avi in {md_folder} -> skip"); return

    frame_times = ffprobe_frame_times(video)
    if frame_times is None:
        print(f"[WARN] {sid}: ffprobe failed on {video.name} -> skip"); return

    stm_path = find_stm_file(md_folder, sid, md_drive)
    if not stm_path:
        print(f"[WARN] {sid}: no STM (.stm/.stm.xlsx) in {md_folder} -> skip"); return

    try:
        intervals = read_intervals_seconds(stm_path)
    except Exception as e:
        print(f"[WARN] {sid}: failed to read intervals from {stm_path.name}: {e} -> skip"); return

    labels = label_frames_from_intervals(frame_times, intervals)

    out_dir.mkdir(parents=True, exist_ok=True)
    pd.DataFrame({
        "frame_idx": np.arange(len(frame_times), dtype=int),
        "t_sec": frame_times,
        "label": labels.astype(int),
    }).to_csv(out_csv, index=False)

    meta = {
        "subject": sid,
        "md_drive": md_drive,
        "video": str(video),
        "stm_file": str(stm_path),
        "n_frames": int(len(frame_times)),
        "n_intervals": int(len(intervals)),
        "intervals_preview": intervals[:5],
        "note": "label=1 if t_sec ∈ any [start,end] (inclusive).",
    }
    meta_json.write_text(json.dumps(meta, indent=2))
    print(f"[OK] {sid}: wrote {out_csv} and {meta_json.name}")

def main():
    ap = argparse.ArgumentParser(description="Per-frame labels from STM intervals (VFR-safe).")
    ap.add_argument("--from", dest="from_id", type=str, default=None,
                    help="Start from subject (e.g., T032).")
    ap.add_argument("--only", dest="only", type=str, default=None,
                    help="Comma-separated subjects to process (e.g., T032,T033).")
    args = ap.parse_args()

    subs = find_subject_dirs(ROOT)
    if not subs:
        print("[ERR] no subject folders found"); return

    if args.only:
        wanted = [s.strip().upper() for s in args.only.split(",") if s.strip()]
        submap = {p.name.upper(): p for p in subs}
        subs = [submap[s] for s in wanted if s in submap]
        if not subs:
            print(f"[ERR] none of requested subjects found: {wanted}"); return
    elif args.from_id:
        start = args.from_id.strip().upper()
        idx = next((i for i, p in enumerate(subs) if p.name.upper() == start), None)
        if idx is None:
            print(f"[ERR] start subject not found: {start}"); return
        subs = subs[idx:]

    for s in subs:
        process_subject(s, OUT_DIR)

if __name__ == "__main__":
    main()
